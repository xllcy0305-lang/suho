# -*- coding: utf-8 -*-
"""Excel 导出模块"""

import logging
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

logger = logging.getLogger(__name__)

HDR_FONT = Font(bold=True, size=11, color="FFFFFF")
HDR_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HDR_ALIGN = Alignment(horizontal="center", vertical="center")


def export_titles_xlsx(titles: list[dict], plat_key: str, product_cn: str,
                       platform_limits: dict) -> tuple[BytesIO, str]:
    plat = platform_limits.get(plat_key, {})
    pn = plat.get("name_cn", plat_key)
    mc = plat.get("max_chars", 200)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SEO标题"

    headers = ["序号", "SEO 标题", "字符数", "AI评分", "平台", "字数上限"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN

    for i, item in enumerate(titles, 2):
        title = item["title"] if isinstance(item, dict) else item
        chars = item.get("chars", len(title)) if isinstance(item, dict) else len(title)
        score = item.get("score", 0) if isinstance(item, dict) else 0
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=title)
        ws.cell(row=i, column=3, value=chars)
        ws.cell(row=i, column=4, value=score)
        ws.cell(row=i, column=5, value=pn)
        ws.cell(row=i, column=6, value=mc)
        # 高分标绿
        if score >= 80:
            ws.cell(row=i, column=4).fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            )

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 8

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f"{pn}_{product_cn}_SEO标题_{ts}.xlsx"

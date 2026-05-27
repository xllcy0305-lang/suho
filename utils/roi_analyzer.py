# -*- coding: utf-8 -*-
"""ROI 分析模块"""

import logging
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006", bold=True)
HDR_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def analyze_roi(uploaded_file) -> tuple[BytesIO | None, str]:
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=False)
    except Exception as e:
        return None, f"读取失败: {e}"

    try:
        for sname in wb.sheetnames:
            ws = wb[sname]
            mc = ws.max_column
            mr = ws.max_row

            # 定位毛利率列
            mcol = None
            for c in range(1, mc + 1):
                v = ws.cell(row=1, column=c).value
                if v and isinstance(v, str):
                    vl = v.lower().strip()
                    if "毛利率" in vl or "margin" in vl or "利润" in vl:
                        mcol = c
                        break

            c1 = mc + 1
            c2 = mc + 2
            for col, hdr in [(c1, "单品保本ROI\n(1/毛利率)"), (c2, "全店通投保本ROI")]:
                cell = ws.cell(row=1, column=col, value=hdr)
                cell.font = HDR_FONT
                cell.fill = HDR_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.column_dimensions[get_column_letter(col)].width = 18

            margins = []
            for r in range(2, mr + 1):
                if mcol:
                    raw = ws.cell(row=r, column=mcol).value
                    try:
                        m = float(raw) if raw is not None else None
                    except (ValueError, TypeError):
                        m = None
                    if m is not None and m > 0:
                        roi = round(1.0 / m, 4)
                        cell = ws.cell(row=r, column=c1, value=roi)
                        cell.number_format = "0.0000"
                        if roi > 2.5:
                            cell.fill = RED_FILL
                            cell.font = RED_FONT
                        elif roi <= 1.5:
                            cell.fill = GREEN_FILL
                        margins.append(m)
                    else:
                        ws.cell(row=r, column=c1, value="N/A")
                else:
                    ws.cell(row=r, column=c1, value="未找到毛利率列")

            if margins:
                avg = sum(margins) / len(margins)
                if avg > 0:
                    sroi = round(1.0 / avg, 4)
                    for r in range(2, mr + 1):
                        cell = ws.cell(row=r, column=c2, value=sroi)
                        cell.number_format = "0.0000"
                        if sroi > 2.5:
                            cell.fill = RED_FILL
                            cell.font = RED_FONT
                        elif sroi <= 1.5:
                            cell.fill = GREEN_FILL

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return buf, f"实际利润表_广告保本ROI分析版_{ts}.xlsx"
    except Exception as e:
        return None, f"分析异常: {e}"

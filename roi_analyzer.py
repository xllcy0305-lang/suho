# -*- coding: utf-8 -*-
"""ROI 分析模块 — 大文件优化版"""

import gc
import logging
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006", bold=True)
HDR_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# 文件大小限制
MAX_FILE_SIZE = 1 * 1024 * 1024 * 1024       # 1GB 硬限制
WARN_FILE_SIZE = 500 * 1024 * 1024            # 500MB 警告线
CHUNK_SIZE = 5000                              # 每批处理行数


def check_file_size(file_bytes: bytes) -> tuple[bool, str, str]:
    """检查文件大小，返回 (是否可处理, 消息, 级别)"""
    size = len(file_bytes)
    size_mb = size / (1024 * 1024)

    if size > MAX_FILE_SIZE:
        return False, f"文件过大 ({size_mb:.1f}MB)，超过 1GB 限制。请拆分后上传。", "error"
    if size > WARN_FILE_SIZE:
        return True, f"文件较大 ({size_mb:.1f}MB)，处理可能需要较长时间，请耐心等待。", "warning"
    if size > 50 * 1024 * 1024:
        return True, f"文件 {size_mb:.1f}MB，正在处理中…", "info"
    return True, f"文件 {size_mb:.1f}MB", "info"


def analyze_roi(uploaded_file, progress_callback=None) -> tuple[BytesIO | None, str]:
    """
    分析 Excel 利润表，追加保本 ROI 列。
    progress_callback(step: str, pct: int) 用于报告进度。
    """
    def _progress(step, pct):
        if progress_callback:
            progress_callback(step, pct)

    try:
        # 读取文件内容到内存（Streamlit uploaded file 需要）
        _progress("读取文件…", 5)
        file_bytes = uploaded_file.read()
        uploaded_file.seek(0)

        # 检查文件大小
        ok, msg, level = check_file_size(file_bytes)
        if not ok:
            return None, msg

        _progress("打开工作簿（只读扫描）…", 10)

        # 第一遍：只读模式扫描结构
        file_buf = BytesIO(file_bytes)
        wb_read = openpyxl.load_workbook(file_buf, read_only=True, data_only=True)
        sheet_info = []

        for sname in wb_read.sheetnames:
            ws = wb_read[sname]
            # 定位毛利率列
            mcol = None
            max_col = 0
            row_count = 0

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    max_col = len(row)
                    for c, v in enumerate(row):
                        if v and isinstance(v, str):
                            vl = v.lower().strip()
                            if "毛利率" in vl or "margin" in vl or "利润" in vl:
                                mcol = c
                                break
                else:
                    row_count += 1

            sheet_info.append({
                "name": sname,
                "mcol": mcol,
                "max_col": max_col,
                "row_count": row_count,
            })

        wb_read.close()
        del wb_read
        gc.collect()

        _progress("只读扫描完成，开始计算…", 25)

        # 第二遍：读写模式，逐块处理
        file_buf.seek(0)
        wb = openpyxl.load_workbook(file_buf)
        total_sheets = len(sheet_info)

        for si, info in enumerate(sheet_info):
            sname = info["name"]
            mcol = info["mcol"]
            mr = info["row_count"] + 1  # +1 for header
            mc = info["max_col"]

            _progress(f"处理工作表 [{sname}] ({si+1}/{total_sheets})…",
                      30 + int(60 * si / max(total_sheets, 1)))

            ws = wb[sname]

            # 写入表头
            c1 = mc + 1
            c2 = mc + 2
            for col, hdr in [(c1, "单品保本ROI\n(1/毛利率)"), (c2, "全店通投保本ROI")]:
                cell = ws.cell(row=1, column=col, value=hdr)
                cell.font = HDR_FONT
                cell.fill = HDR_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.column_dimensions[get_column_letter(col)].width = 18

            # 逐块处理数据行
            margins = []
            for chunk_start in range(2, mr + 1, CHUNK_SIZE):
                chunk_end = min(chunk_start + CHUNK_SIZE, mr + 1)
                for r in range(chunk_start, chunk_end):
                    if mcol is not None:
                        raw = ws.cell(row=r, column=mcol + 1).value  # openpyxl 1-indexed
                        try:
                            m = float(raw) if raw is not None else None
                        except (ValueError, TypeError):
                            m = None
                        if m is not None and m > 0:
                            roi = round(1.0 / m, 4)
                            cell = ws.cell(row=r, column=c1, value=roi)
                            cell.number_format = "0.0000"
                            if roi > 2.5:
                                cell.fill = RED_FILL
                                cell.font = RED_FONT
                            elif roi <= 1.5:
                                cell.fill = GREEN_FILL
                            margins.append(m)
                        else:
                            ws.cell(row=r, column=c1, value="N/A" if raw is None else "无效")
                    else:
                        ws.cell(row=r, column=c1, value="未找到毛利率列")

                # 每个 chunk 后释放内存
                gc.collect()

            # 写入全店通投 ROI
            if margins:
                avg = sum(margins) / len(margins)
                if avg > 0:
                    sroi = round(1.0 / avg, 4)
                    for r in range(2, mr + 1):
                        cell = ws.cell(row=r, column=c2, value=sroi)
                        cell.number_format = "0.0000"
                        if sroi > 2.5:
                            cell.fill = RED_FILL
                            cell.font = RED_FONT
                        elif sroi <= 1.5:
                            cell.fill = GREEN_FILL

        _progress("保存结果…", 95)
        buf = BytesIO()
        wb.save(buf)
        wb.close()
        del wb
        gc.collect()

        buf.seek(0)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        _progress("完成！", 100)
        return buf, f"实际利润表_广告保本ROI分析版_{ts}.xlsx"

    except MemoryError:
        gc.collect()
        return None, "内存不足，请拆分文件后重试。建议单文件不超过 500MB。"
    except Exception as e:
        gc.collect()
        return None, f"分析异常: {e}"
